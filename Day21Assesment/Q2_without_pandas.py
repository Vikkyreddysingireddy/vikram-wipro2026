from openpyxl import load_workbook, Workbook

wb = load_workbook("sales_data.xlsx")
ws = wb["2025"]

new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "Summary"

headers = [cell.value for cell in ws[1]]
headers.append("Total")
new_ws.append(headers)

for row in ws.iter_rows(min_row=2, values_only=True):
    product, quantity, price = row
    total = quantity * price
    new_ws.append([product, quantity, price, total])

new_wb.save("sales_summary.xlsx")
print("sales_summary.xlsx created using OpenPyXL")
